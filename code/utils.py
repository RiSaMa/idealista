import os
import base64
import requests
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment

verify = True # when running in macos locally, it might have to be False due to some issue

# Get secrets
api_key = os.getenv('IDEALISTA_API_KEY')
secret = os.getenv('IDEALISTA_API_SECRET')

def get_oauth_token():
    # URL encode the API key and secret
    credentials = f"{api_key}:{secret}"
    
    # Base64 encode the credentials
    encoded_credentials = base64.b64encode(credentials.encode('utf-8')).decode('utf-8')
    
    # Set up the request headers
    headers = {
        'Authorization': f'Basic {encoded_credentials}',
        'Content-Type': 'application/x-www-form-urlencoded;charset=UTF-8'
    }
    
    # Set up the request body
    data = {
        'grant_type': 'client_credentials',
        'scope': 'read'
    }
    
    # Make the POST request to get the token
    response = requests.post('https://api.idealista.com/oauth/token', headers=headers, data=data, verify=verify)
    
    # Check if the request was successful
    if response.status_code == 200:
        print("Authentication correct.")
        return response.json().get('access_token')
    else:
        # Print the error and return None
        print("Authentication failed.")
        print(f"Error: {response.status_code} - {response.text}")
        return None

def search_properties(page_num, since_date):

    # Get the OAuth token
    token = get_oauth_token()

    # Define the endpoint URL
    url = "https://api.idealista.com/3.5/es/search"

    # Define the search parameters
    params = {
        "country": "es",  # Spain
        "operation": "sale",
        "propertyType": "homes",
        "center": "40.44672907846094,-3.692355207550619",  # Approximate coordinates for center of Madrid (Nuevos Ministerios)
        "distance": 25000,  # meters
        "maxItems": 50,
        "numPage": page_num,
        "maxPrice": 150001,
        "minPrice": 90000,
        "sinceDate": since_date,  # M: last month, for initial search. W: last week for recurrent searches
        "order": "price",
        "sort": "asc",
        "bedrooms": "2",  # At least 2 bedrooms
        "flat": True,
        "minSize": 45  # At least 45 m²
    }

    # Define the headers, including the authorization token
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/x-www-form-urlencoded;charset=UTF-8"
    }

    # Make the API request
    response = requests.post(url, headers=headers, data=params, verify=verify)

    # Check if the request was successful
    if response.status_code == 200:
        # Parse and return the JSON response
        return response.json().get('elementList', [])
    else:
        # Print the error and return an empty list
        print(f"Error: {response.status_code} - {response.text}")
        return []

def filter_properties(properties):
    # Words to filter out properties
    exclude_keywords = ["nuda", "alquilado", "alquilada", "no se puede", "ocupado", "okupado", "subasta", "ilegal"]
    filtered_properties = []

    for property in properties:
        # Check if description contains any exclusion keywords
        if 'description' not in property or any(re.search(keyword, property['description'], re.IGNORECASE) for keyword in exclude_keywords):
            continue

        # Check if the property has at least 3 pictures
        if property['numPhotos'] < 3:
            continue
        
        if "propertyCode" not in property or "url" not in property:
            continue
        
        # Add property to the filtered list
        filtered_properties.append({
            'propertyCode': property['propertyCode'],
            'url': property['url'],
            'price': property.get('price', 'N/A'),
            'size': property.get('size', 'N/A'),
            'address': property.get('address', 'N/A'),
            'bedrooms': property.get('rooms', 'N/A'),
            'floor': property.get('floor', 'N/A'),
            'description': property.get('description', 'NA'),
            'Interested?': '',  # Default empty --  needs manual check
            'Contacted?': ''  # Default empty
        })

    return filtered_properties


def update_database(new_properties, local_file_path):
    # Load existing data with pandas
    try:
        df_existing = pd.read_excel(local_file_path)
    except FileNotFoundError:
        # Create an empty DataFrame if the file doesn't exist
        df_existing = pd.DataFrame(columns=['propertyCode', 'url', 'price', 'size', 'address', 'bedrooms', 'floor', 'description', 'Interested?', 'Contacted?'])

    # Convert new properties to DataFrame
    df_new = pd.DataFrame(new_properties)

    # Ensure 'propertyCode' is of the same type in both DataFrames
    df_existing['propertyCode'] = df_existing['propertyCode'].astype(str)
    df_new['propertyCode'] = df_new['propertyCode'].astype(str)

    # Merge new properties with existing database
    df_updated = pd.concat([df_existing, df_new], ignore_index=True).drop_duplicates(subset='propertyCode', keep='first')

    # Save updated data to Excel
    df_updated.to_excel(local_file_path, index=False)

    number_new_flats = len(df_updated)-len(df_existing)

    # Load the workbook with openpyxl
    wb = load_workbook(local_file_path)
    ws = wb.active

    # Format the range as a table
    tab = Table(displayName="PropertiesTable", ref=ws.dimensions)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Make URLs clickable hyperlinks
    for cell in ws['B'][1:]:  # Assuming 'B' is the URL column
        cell.hyperlink = cell.value
        cell.value = "Link"
        cell.style = "Hyperlink"

        
    # Adjust column widths
    for col in ws.columns:
        column = col[0].column_letter  # Get the column name
        if column in ['A', 'B', 'D', 'F', 'G']:
            width = 10
        elif column=='C':
            # Apply euro currency format
            for cell in col[1:]:  # Skip header
                cell.number_format = '€#,##0'  # Custom format for euros without decimals
        elif column == 'E':  # ADDRESS column
            width = 30
        elif column == 'H':  # Description column
            width = 120
            #for cell in col:
            #    cell.alignment = Alignment(wrap_text=True)  # Wrap text in description cells
        elif column in ['I', 'J']:
            width = 15

        # Apply vertical alignment to all cells
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

        ws.column_dimensions[column].width = width

    # Save the formatted workbook to the original file
    wb.save(local_file_path)

    return number_new_flats

